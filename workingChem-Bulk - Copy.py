import openpyxl
import datetime
import os
import sys



def screening():
    print('**screening sequences')
###remove all but chem and DNU

###remove all none Bulk
    print('assembling sheet 2')
    for i in range(1, sheet.max_row):
        x=sheet.cell(row=i, column=3).value
        if x[:4]=="CHEM" or x[:3]=="DNU":
            list = (sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value, sheet.cell(row=i, column=3).value)
            sheet2.append(list)

    print('assembling sheet 3')
    for j in range(1, sheet2.max_row):
        y=sheet2.cell(row=j, column=3).value
        if y[:-5:-1] =="KLUB":
            lastvalue = sheet2.cell(row=j, column=3).value
            shortvalue = lastvalue[5:-5]
            splitvalue = shortvalue.split(",")
            truncated = splitvalue[0]
            list = (sheet2.cell(row=j, column=1).value, sheet2.cell(row=j, column=2).value, truncated)
            sheet3.append(list)

    print('sorting list')
    for k in range(1, sheet3.max_row+1):
        rawsapnumber.append(sheet3.cell(row=k , column=1).value)
    rawsapnumber.sort()
    print(len(rawsapnumber))

    l=0
    while l < len(rawsapnumber):
        if rawsapnumber[l] == rawsapnumber[l + 1]:
            rawsapnumber.pop(l+1)
        l = l+1
    print(len(rawsapnumber))

    for m in range(0, len(rawsapnumber)):
        lookup2=str(m+1)
        lookup3=str(sheet3.max_row)
        list = (rawsapnumber[m],"=vlookup(A" + lookup2 + ",sheet3!A1:C" + lookup3 + ",2,False)","=vlookup(A" + lookup2 + ",sheet3!A1:C" + lookup3 + ",3,False)"  )
        sheet4.append(list)


    finish()


def finish():
    del wb['Sheet2']
    del wb['Sheet1']
    now = datetime.datetime.now()
    timestamp = str(now.strftime("%Y%m%d_%H%M%S"))

    print ("***Creating New File")
    filepath="C:\\Users\\HB80464\\Desktop\\Python\\"

    wb.save(filepath+timestamp +'.xlsx')
    wb.close()


try:
   droppedFile = input("drop file")
except IndexError:
    print("No file dropped")

print("**Opening" , droppedFile)

wb = openpyxl.load_workbook(droppedFile)

while len(wb.sheetnames)>1:
    del wb[wb.sheetnames[1]]

wb.sheetnames[0]="Sheet1"
sheet= wb['Sheet1']
wb.create_sheet('Sheet2')
wb.create_sheet('Sheet3')
wb.create_sheet('Sheet4')
sheet2=wb['Sheet2']
sheet3=wb['Sheet3']
sheet4=wb['Sheet4']
list = ()
directory=[]
rawsapnumber = []

print("**deleting")

sheet.delete_rows(1,3)
sheet.delete_cols(1,2)
sheet.delete_cols(4,sheet.max_column)

screening()
